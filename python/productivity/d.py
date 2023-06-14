from openpyxl import load_workbook
from pathlib import Path,PurePath
from openpyxl import Workbook

src_path='/Users/tx/Desktop/文章1代码/调查问卷'
dst_path = '/Users/tx/Desktop/文章1代码/result/结果.xlsx'
p= Path(src_path)
files= [x for x in p.iterdir() if PurePath(x).match('*.xlsx')]
# print(files)
content = []
for file in files:
    username = file.stem
    # print(file)
    wb = load_workbook(file)
    # print(wb.sheetnames)
    ws=wb.active
    # print(ws)
    answer1 = ws['E5'].value
    answer2 = ws.cell(row=11,column=5).value
    print(answer1)
    print(answer2)
    temp = f'{username},{answer1},{answer2}'
    content.append(temp.split(','))
    print (temp)
    print(content)

ws1_header =['员工姓名','第一题','第二题']
wb1= Workbook()
ws1 = wb1.active
ws1 = wb1.create_sheet("统计结果")
rw=1
col=1
for cell_header in ws1_header:
     ws1.cell(row=rw, column=col,value=cell_header)
     col +=1

rw +=1
for line in content:
    col=1
    for cell in line:
        ws1.cell(row=rw, column=col, value=cell)
        col+=1
    rw+=1

wb1.save(dst_path)