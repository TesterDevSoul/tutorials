# 调查表单的读写
import os
from io import BytesIO

import openpyxl
import pandas as pd
from openpyxl.drawing.image import Image

import xlrd
from pathlib import Path, PurePath

import xlwt
from matplotlib import pyplot as plt

print("----------------  获取 Sheet2 调查表结果  -------------")

# file = '/Users/gaigai/Desktop/测试用例/LiteMall后台管理系统商品查询接口用例.xls'

# 指定要合并excel的路径
src_path = '/Users/gaigai/encoding/zh-cn/编程语言/Python/自动化办公/Excel操作/xls文件/用例问卷'


# 取得该目录下所有的xls格式文件
# /Users/gaigai/Desktop/测试用例
p = Path(src_path)
# 列表推导式
files = [x for x in p.iterdir() if PurePath(x).match('*.xls')]
# files = []
# p.iterdir() #遍历目录 p 中的所有文件和子目录
# for x in p.iterdir():
# # 判断文件路径 x 是否匹配 '*.xls' 模式，即文件扩展名是否为 `.xls`
#     if PurePath(x).match('*.xls'):
#         files.append(x)

# 结果数据存储
content = []
# 标题命名
table_header = ['员工姓名', '第一题', '第二题']
# 结果数据添加标题
# content.append(table_header)

# todo: 双层for循环读取文件夹内的每一行
# 对每一个文件进行重复处理
for file in files:

    # 使用了os库的os.path.getsize()函数来检查Excel文件的大小，
    # 如果文件大小大于0，将继续读取并处理该文件。
    # 将输出一个错误消息。这将避免抛出XLRDError: File size is 0 bytes错误。
    file_size = os.path.getsize(file)
    # 如果文件大小大于0，则继续读取文件
    if file_size > 0:
        # 如果文件大小大于0，则继续读取文件
        # 打开 Excel 文件，file 是文件路径或类似文件对象的参数
        data = xlrd.open_workbook(file)
        # sheet2的选项结果
        # username = "韩梅梅"
        # 1. 根据下标获取sheet
        sheet2 = data.sheet_by_index(1)
        # 2. 获取sheet的名称 作为员工姓名 打印
        username = sheet2.name
        print(username)
        # sheet2 = data.sheet_by_name(username)
        # 取得每一项的结果
        # 3. 获取第一道题的选项
        answer1 = sheet2.cell_value(rowx=4, colx=4)
        # 4. 获取第二道题的选项
        answer2 = sheet2.cell_value(rowx=10, colx=4)
        # 5. 合并为一行存储
        temp = f'{username},{answer1},{answer2}'
        # 合并为一行先存储起来
        content.append(temp.split(','))
        print(content)

print("----------------  写入文件 Excel 文件中的所有行  -------------")

workbook = xlwt.Workbook(encoding='utf-8')
worksheet = workbook.add_sheet("调查结果")

for col_num, column_title in enumerate(table_header):
    # 第一行写入列标题 0 第一行 col_num
    worksheet.write(0, col_num, column_title)

# dict_rows 是一个嵌套列表，而不是字典列表
# 通过 enumerate() 函数同时遍历 dict_rows（一个字典的列表）的索引（行号）和元素（行数据）。
# start=1 参数表示下标从 1 开始计数，而不是默认从 0 开始。
for row_num, row_data in enumerate(content, start=1):
    for col_num, cell_value in enumerate(row_data):
        # 将列表中的数据写入对应的单元格
        worksheet.write(row_num, col_num, cell_value)

# 保存 Excel 文件
# 指定合并完成的路径
dst_file = '/Users/gaigai/Desktop/调查表汇总.xls'
# new_file_path = 'path/to/new_file.xls'
# workbook.save(new_file_path)
workbook.save(dst_file)


df = pd.read_excel(dst_file, engine="xlrd")

# 另存为新的 Excel 文件（.xlsx）
new_excel = '/Users/gaigai/Desktop/调查表汇总.xlsx'
df.to_excel(new_excel, index=False)


# 加载工作簿并选择工作表
workbook = openpyxl.load_workbook(new_excel)
worksheet = workbook.active
# 设置要检查的列
column_index = 1  # 例如，设置为第一列

# 初始化计数器
A_count = 0
B_count = 0
C_count = 0
D_count = 0

# 遍历列中的所有非空单元格
for cell in worksheet[worksheet.get_column_letter(column_index)][1:]:
    if cell.value == "A":
        A_count += 1
    elif cell.value == "B":
        B_count += 1
    elif cell.value == "C":
        C_count += 1
    elif cell.value == "D":
        D_count += 1
print("A_count count:", A_count)
print("B_count count:", B_count)
print("C_count count:", C_count)
print("D_count count:", D_count)


# 使用 openpyxl 处理新文件并插入图片
# workbook1 = openpyxl.load_workbook(new_excel)
# labels = ['A', 'B', 'C', 'D']
# sizes = [25, 30, 20, 25]
#
# plt.pie(sizes, labels=labels, autopct='%1.1f%%')
# plt.title("title",)
# plt.axis('equal')  # 设置为正圆形
# # plt.show()
# # 将饼图保存到内存中的图片文件中
# buf = BytesIO()
# plt.savefig(buf, format='png')
# buf.seek(0)  # 重置文件指针
#
#
# ws = workbook1.active
#
# image = Image(buf)
# ws.add_image(image, 'D1')  # 将图片插入到 Excel 的 'D1' 单元格
#
# # 保存 Excel 文件
# # workbook1.save(dst_file)
#
#  # 保存含有图像的 Excel 文件
# workbook1.save(new_excel)