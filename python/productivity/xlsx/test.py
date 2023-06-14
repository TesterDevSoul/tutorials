from pathlib import Path, PurePath

import openpyxl

# 指定要合并excel的路径
src_path = '/Users/gaigai/Desktop/测试用例xlsx'
dst_path = '/Users/gaigai/Desktop/结果'


# 取得该目录下所有的xls格式文件
p = Path(src_path)
# 列表推导式
xlsx_files = [x for x in p.iterdir() if PurePath(x).match('*.xlsx')]

# 在内存中创建一个新的空白Excel工作簿
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active
# sheet名 测试用例
output_sheet.title = "测试用例"

# 逐个读取并合并每个Excel文件
for file in xlsx_files:
    # 使用 openpyxl 库的 load_workbook() 函数来加载 Excel 工作簿
    input_workbook = openpyxl.load_workbook(file)
    # 获取活动工作表 工作簿（workbook）中访问活动工作表（active sheet）
    input_sheet = input_workbook.active

    # 逐行和逐列复制输入工作表中的数据
    if file == xlsx_files[0]:  # 如果是第一个文件, 包括表头
        for row in input_sheet.iter_rows():
            row_values = [cell.value for cell in row]
            output_sheet.append(row_values)
    else:  # 其他文件仅复制数据部分，跳过表头
        for row in input_sheet.iter_rows(min_row=2):  # 跳过表头
            row_values = [cell.value for cell in row]
            output_sheet.append(row_values)

    # 关闭输入工作簿
    input_workbook.close()

# 保存合并后的数据到一个新的Excel文件
# output_file = "merged_file.xlsx"
folder = Path(dst_path)
file = 'Little.xlsx'
# 使用 pathlib 进行路径拼接
new_excel = folder / file

output_workbook.save(new_excel)

# 关闭输出工作簿
output_workbook.close()

print(f"合并完成，新文件已保存为：{new_excel}")