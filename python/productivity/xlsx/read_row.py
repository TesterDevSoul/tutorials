
import openpyxl

file = '/Users/gaigai/Desktop/LiteMall.xlsx'


# 使用 openpyxl 库的 load_workbook() 函数来加载 Excel 工作簿
input_workbook = openpyxl.load_workbook(file)
# 获取活动工作表 工作簿（workbook）中访问活动工作表（active sheet）
input_sheet = input_workbook.active

print("----------------  开始遍历 Excel 文件中的所有行  -------------")

# iter_rows()方法允许您遍历工作表中的所有行。
# 遍历了input_sheet中的所有行，并使用row_values列表存储行中的所有单元格值。
for row in input_sheet.iter_rows():
    # 列表推导式
    row_values = [cell.value for cell in row]
    # 初始化一个空列表来存储行中的所有单元格值
    # row_values = []
    # 遍历行中的所有单元格
    # for cell in row:
    # 将单元格值添加到 row_values 列表中
    #     row_values.append(cell.value)
    print(row_values)